
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import pandas as pd
import sys
import os
import configparser

def format(file,indices='auto'):
    def format_3x2(ignore_sheets='Sheet1'):
        pass

    def format_2x2(ignore_sheets='Sheet1'):
        pass

    config = configparser.ConfigParser()
    config.read('excel_style_formats.ini')

    # CONFIG PARAMETERS

    # numeric formats
    nb_format = config['data columns formatting']['format']
    
    # overall grandtotals formats
    current_section = config['overall grand totals formatting']
    overall_grandtotal_fill=PatternFill(start_color=current_section['fill_color'],
        fill_type=current_section['fill_type'])
    overall_grandtotal_font=Font(bold=bool(current_section['bold']),
        size=current_section['size'],
        color=current_section['color'])

    # grandtotals formats
    current_section = config['grand totals formatting']
    grandtotal_fill=PatternFill(start_color=current_section['fill_color'],
        fill_type=current_section['fill_type'])
    grandtotal_font=Font(bold=bool(current_section['bold']),
        size=current_section['size'],
        color=current_section['color'])

    # subtotals formats
    current_section = config['sub totals formatting']
    subtotal_fill=PatternFill(start_color=current_section['fill_color'],
        fill_type=current_section['fill_type'])
    subtotal_font=Font(bold=bool(current_section['bold']),
        size=current_section['size'],
        color=current_section['color'])

    # column widths
    data_width = int(config['widths']['data'])
    col1_width = int(config['widths']['col1'])
    col2_width = int(config['widths']['col2'])
    col3_width = int(config['widths']['col3'])

    # borders
    bd_thin = Side(style='thin', color='000000')

    # headers
    current_section = config['headers formatting']
    header_alignment = Alignment(wrapText=current_section['wrap_text'],
        horizontal=current_section['horizontal_align'],
        vertical=current_section['vertical_align'])
    header_font = Font(size=current_section['size'],bold=bool(current_section['bold']),color=current_section['color'])
    header_fill = PatternFill(start_color=current_section['fill'],fill_type=current_section['fill_type'])

    # sub headers
    current_section = config['sub headers formatting']
    sub_header_alignment = Alignment(wrapText=current_section['wrap_text'],
        horizontal=current_section['horizontal_align'],
        vertical=current_section['vertical_align'])
    sub_header_font = Font(size=current_section['size'],bold=bool(current_section['bold']),color=current_section['color'])
    sub_header_fill = PatternFill(start_color=current_section['fill'],fill_type=current_section['fill_type'])

    alphabet = [0,'A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU']

    wb = load_workbook(filename = 'Final Forecasts_FY2020.xlsx')

    for sheet in wb.sheetnames:
        current_sheet = wb[sheet]
        # ignore raw data sheet
        if sheet == 'raw':
            continue

        if sheet in ['Final Forecast','Original SARIMA Predictions']:
            indices = (2,2)
            min_row = 4
        else:
            indices = (2,1)
            min_row = 2

        min_col = 3
        max_row = 1000

        # SET INDEX COLUMN WIDTHS
        current_sheet.column_dimensions[alphabet[1]].width = col1_width
        current_sheet.column_dimensions[alphabet[2]].width = col2_width


        # FIND NUMBER OF NUMERIC COLUMNS
        column_dict = {} # dictionary where key is numeric position of first numeric column, value is a list: [alphabetical representation of column position, desired format]
        c=0 # count through the columns and add 1 to align the indices
        for column in current_sheet.iter_cols(min_row=min_row,
                                       max_row=min_row, # only searching through the first row that has numeric data (non-column rows)
                                       min_col=min_col, # start where we know the row indices ends
                                       max_col=(len(alphabet)-1)): # end where the alphabet list ends until we know the actual number of columns
            if column[0].value != None:
                column_dict[min_col+c] = [alphabet[min_col+c],fy_format]
                c+=1
            else:
                column_dict[min_col+c-1] = [alphabet[min_col+c-1],yoy_format]
                break # stop after we found the end of the excel sheet

        # FORMAT NUMERIC CELLS
        c = min_col # begin with first numeric column
        max_col = list(column_dict.keys())[-1] # set flag where the columns end
        for column in current_sheet.iter_cols(min_row=min_row,
                               max_row=max_row,
                               min_col=min_col,
                               max_col=max_col):                     
            for i in range(max_row-min_row):
                if column[i].value != None:
                    column[i].number_format = column_dict[c][1] # format each numeric value appropriately
                else:
                    break
            c+=1

        # BOLD AND COLOR ROWS WITH TOTALS
        r=min_row # iterate through rows
        max_row = i+min_row-1 # set flag where last row is in the sheet
        # this loop should accurately represent the actual sheet dimensions now that they are known
        for row in current_sheet.iter_rows(min_row=min_row,
                               max_row=max_row,
                               min_col=min_col-1,
                               max_col=max_col):
            if row[0].value == 'Total':
                if sheet in ['Final Forecast','Original SARIMA Predictions']:
                    # sub totals
                    for i in range(max_col-min_col+1):
                        row[i+1].font = subtotal_font
                        row[i+1].fill = subtotal_fill
     
            if r == max_row:
                if sheet in ['Final Forecast','Original SARIMA Predictions']:
                    # totals
                    for i in range(max_col-min_col+1):
                        row[i+1].font = overall_grandtotal_font
                        row[i+1].fill = overall_grandtotal_fill
                for i in range(max_col-min_col+1):
                    row[i+1].border = Border(bottom=bd_thin)
            r+=1

        # ADD BORDERS AROUND EVERY NUMERIC COLUMN
        c = 1
        for column in current_sheet.iter_cols(min_row=1,
                               max_row=max_row,
                               min_col=min_col,
                               max_col=max_col):
            for r in range(max_row):
                if r == indices[1]-1:
                    column[r].border = Border(bottom=bd_thin,top=bd_thin,left=bd_thin,right=bd_thin) 
                elif r < max_row-1:
                    column[r].border = Border(right=bd_thin)
                else:
                    column[r].border = Border(bottom=bd_thin,right=bd_thin) 
            c+=1

        # SPACING
        for i in column_dict:
            current_sheet.column_dimensions[column_dict[i][0]].width = data_width

        # HEADER FORMATS
        for column in current_sheet.iter_cols(min_row=1,
                               max_row=1,
                               min_col=1,
                               max_col=max_col):
            column[0].alignment = header_alignment
            column[0].font = header_font
            column[0].fill = header_fill

        # SUB HEADER FORMATS
        if indices[1] > 1:
            for column in current_sheet.iter_cols(min_row=2,
                                   max_row=2,
                                   min_col=min_col,
                                   max_col=max_col):
                column[0].alignment = sub_header_alignment
                column[0].font = sub_header_font
                column[0].fill = sub_header_fill

        # MERGE ROWS
        for i in range(indices[0]):
            if current_sheet['{}{}'.format(alphabet[i+1],min_row-1)].value != None:
                current_sheet['{}{}'.format(alphabet[i+1],1)].value = current_sheet['{}{}'.format(alphabet[i+1],min_row-1)].value
                current_sheet.merge_cells('{}{}:{}{}'.format(alphabet[i+1],1,alphabet[i+1],min_row-1))

        current_sheet['A1'].value = 'LOB'
        if sheet in ['Final Forecast','Original SARIMA Predictions']:
            current_sheet['B1'].value = 'Store'
        else:
            current_sheet['B1'].value = 'Category'

        # ADD BORDERS AROUND THE INDEX COLUMNS
        c = 1
        for column in current_sheet.iter_cols(min_row=1,
                               max_row=max_row,
                               min_col=1,
                               max_col=min_col-1):
            for r in range(max_row):
                column[r].border = Border(bottom=bd_thin,top=bd_thin,left=bd_thin,right=bd_thin)          

    wb.save('Final Forecasts_FY2020.xlsx')

if __name__ == '__main__':
    main()

def format_2x1(pivot_tbl):
	pass
def format_2x2(pivot_tbl):
	pass
def format_3x1(pivot_tbl):
	pass
def format_3x2(pivot_tbl):
	pass